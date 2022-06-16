import json
from typing import List
import toml

import logging
import requests
import datetime
import xlsxwriter
from openpyxl import load_workbook

# aavid = 1733053681002496

cookie = 'MONITOR_WEB_ID=ba14d717-dbea-45fd-8f60-94acdcbc5f88; ttcid=2833c61bbf894df7ba034710ea192f0d36; qc_tt_tag=0; _tea_utm_cache_1574={"utm_source":"qianchuan-origin-entrance","utm_medium":"baiying-pc","utm_campaign":"author","utm_term":"qianchuan-livepromote"}; passport_csrf_token=0369db0e8030c5e25787879a32170569; passport_csrf_token_default=0369db0e8030c5e25787879a32170569; d_ticket=909082e5a3b2bb2b452a8c8a6da1020dd2fad; n_mh=-mqxKVrAml8wYjvJQK_ciUbaL4EURXJiZQ9FV6AtQec; passport_auth_status=d664a4b60cba8e68d2192ab90df979ed,; passport_auth_status_ss=d664a4b60cba8e68d2192ab90df979ed,; sso_auth_status=e0a754c8e17fae23345ceb1070cd506d; sso_auth_status_ss=e0a754c8e17fae23345ceb1070cd506d; ucas_c0=CkAKBTEuMC4wEKmIher-y7jAYhjmJiCv8cCGoY3AAyiwITCcp6DN3oxpQPzEg5QGSPz4v5YGUIC81vLVw5yvYlhvEhR2oEcJ7mhsaX2Gs53DehsUxtHJ-w; ucas_c0_ss=CkAKBTEuMC4wEKmIher-y7jAYhjmJiCv8cCGoY3AAyiwITCcp6DN3oxpQPzEg5QGSPz4v5YGUIC81vLVw5yvYlhvEhR2oEcJ7mhsaX2Gs53DehsUxtHJ-w; gftoken=YmUyMDM0MTZiMHwxNjUyNjIyNzcyMTR8fDAGBgYGBgY; MONITOR_DEVICE_ID=df730b0f-15b3-4430-a199-4eaf7e83cc20; ttwid=1|T9h3daLJFRlIkAOdneZSB5fUHVorG5mPpIK61BtAg6c|1654526332|e241dd2b7047364ac3773786ad0144cb5c6922dbedf6d986338ee55d39d1f7c3; msToken=Oe4Kh59KtYhq8DvahzbdGm5TTmphyFbNjAOTjTT6j3kk_wfdM92ZWDplJZ2qR55CyUg-pN7yO4M5LIPy2Ov03D3bfONoqcAxhvyetXybmaU-ao2PvfqSeLWCgbEYbvY=; x-jupiter-uuid=16548721719201304; _tea_utm_cache_4499=undefined; _ssa_username=undefined; csrftoken=MxT1w068-wtaQYGG9nzarT3vfnrHfLOlAOBo; _ssa_userphone=; tt_scid=DzrSSoq7H616Ymxpv04DZFLwwyeKFe7iIsf5chyvCRx9.Nk9hhRFUInndDqfqQ3P2242; acsessionid=762593c2248c46939c81935de46d6e29'

headers = {
    "cookie": cookie,
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.55 Safari/537.36 Edg/96.0.1054.43",
    "x-csrftoken": "MxT1w068-wtaQYGG9nzarT3vfnrHfLOlAOBo",
    "Content-Type": "application/json",
}


def get_today_str():
    return datetime.datetime.today().strftime("%Y-%m-%d")


def get_current_time_info():
    year = datetime.datetime.now().today().year
    month = datetime.datetime.now().today().month
    day = datetime.datetime.now().today().day
    hour = datetime.datetime.now().time().hour
    minute = datetime.datetime.now().time().minute

    return year, month, day, hour, minute


def get_last_hour_time():
    now_time = datetime.datetime.now().replace(minute=0, second=0, microsecond=0)
    last_hour_time = now_time+datetime.timedelta(hours=-1)
    return last_hour_time

def get_last_hour_time_info():
    last_hour = get_last_hour_time()
    return last_hour.year, last_hour.month, last_hour.day, last_hour.hour


def get_url_data(aavid, today: str, user_info: dict, time_info_timestamp: int) -> dict:
    print(f"get current {aavid=}'s data, please wating...\n")
    logging.info(f"get current {aavid=}'s data, please wating...\n")
    dt = {}



    data = {
        "adFilter": {
            "pricingCategory": 2,
            "app": 0,
            "marGoal": 10,
            "externalAction": [],
            "optimizeGoal": [],
        },
        "creativeFilter": {"app": 0, "marGoal": 10},
        "statsParameter": {
            "startTime": today,
            "endTime": today,
            "metrics": [
                "stat_cost",
                "show_cnt",
                "click_cnt",
                "ctr",
                "cpm",
                "convert_rate",
                "luban_live_enter_cnt",
                "live_watch_one_minute_count",
                "luban_live_slidecart_click_cnt",
                "luban_live_click_product_cnt",
                "luban_live_comment_cnt",
                "luban_live_share_cnt",
                "dy_comment",
                "dy_follow",
                "pay_order_amount",
                "pay_order_count",
                "indirect_order_pay_count_7days",
                "indirect_order_pay_gmv_7days",
                "convert_cnt",
                "convert_cost",
            ],
            "timeDimension": "stat_time_hour",
            "mainDimension": "advertiser_id",
            "pageParams": {"page": -1, "pageSize": 0},
        },
        "aavid": aavid,
    }
    _cook = user_info.get("cookie")
    _cf = user_info.get("carton")

    header = headers
    header["cookie"] = _cook
    header["x-csrftoken"] = _cf

    json_obj = json.dumps(data)

    url = f"https://qianchuan.jinritemai.com/ad/marketing/data/api/v1/report/stats?aavid={aavid}&gfversion=1.0.0.8188"
    resp = requests.post(url=url, headers=header, data=json_obj)
    if resp.status_code >= 400:
        logging.error(f"can not receive {aavid=} data， may be u should change cookie & csrftoken\n")
        print(f"can not receive {aavid=} data, may be u should change cookie & csrftoken\n")
        return dt

    all_data = resp.json()
    if all_data.get("status_code") != 0:
        logging.error(f"can not receive {aavid=} data, may be u should change pay_load or change cookie & csrftoken\n")
        print(f"can not receive {aavid=} data, may be u should change pay_load or change cookie & csrftoken\n")
        return dt

    result = all_data.get("data").get("data")

    stats_data_list = result.get("statsDataRows")
    for stats_data in stats_data_list:
        user_info = stats_data.get("dimensions")
        stat_time_hour = user_info.get("statTimeHour")

        # 拿取 上一个小时 - 当前整点的数据
        if stat_time_hour == str(time_info_timestamp):
            logging.info("can get data at current time and will be write data to file :-)\n")
            need_data = stats_data.get("metrics")

            # 消耗
            cost = need_data.get("cost").get("value") if need_data.get("cost") else 0
            # 展示次数
            show_cnt = (
                need_data.get("showCnt").get("value") if need_data.get("showCnt") else 0
            )
            # 点击次数
            click_cnt = (
                need_data.get("clickCnt").get("value")
                if need_data.get("clickCnt")
                else 0
            )
            # 直播间超过1分钟观看人次
            more_1_minute_cnt = (
                need_data.get("liveWatchOneMinuteCount").get("value")
                if need_data.get("liveWatchOneMinuteCount")
                else 0
            )
            # 直播间查看购物车次数
            sidecar_click_cnt = (
                need_data.get("lubanLiveSlidecartClickCnt").get("value")
                if need_data.get("lubanLiveSlidecartClickCnt")
                else 0
            )
            # 直播间商品点击次数
            product_click_cnt = (
                need_data.get("lubanLiveClickProductCnt").get("value")
                if need_data.get("lubanLiveClickProductCnt")
                else 0
            )
            # 直接成交金额(元)
            direct_order_pay_gmv = (
                need_data.get("directOrderPayGmv").get("value")
                if need_data.get("directOrderPayGmv")
                else 0
            )
            # 直接成交订单数
            direct_order_pay_count = (
                need_data.get("directOrderPayCount").get("value")
                if need_data.get("directOrderPayCount")
                else 0
            )

            dt["cost"] = cost
            dt["show_cnt"] = show_cnt
            dt["click_cnt"] = click_cnt
            dt["more_1_minute_cnt"] = more_1_minute_cnt
            dt["sidecar_click_cnt"] = sidecar_click_cnt
            dt["product_click_cnt"] = product_click_cnt
            dt["direct_order_pay_gmv"] = direct_order_pay_gmv
            dt["direct_order_pay_count"] = direct_order_pay_count

            return dt
    return dt


# 目前此函数并未使用
def get_all_data():
    data_list: List[dict] = []
    today = get_today_str()
    for name, aavid in accountID_dict.items():
        dt = get_url_data(aavid=aavid, today=today)
        dt["name"] = name
        data_list.append(dt)
    return data_list


# 生成 excel 文件
def gen_xlsx():
    _, month, day, hour = get_last_hour_time_info()

    file = f"./output_file/{month}月{day}日{hour}时-{hour+1}时数据.xlsx"

    workbook = xlsxwriter.Workbook(file)
    worksheet = workbook.add_worksheet()

    worksheet.write("A1", "日期")
    worksheet.write("B1", "时间")
    worksheet.write("C1", "账号名称")

    worksheet.write("D1", "消耗(元)")
    worksheet.write("E1", "展示次数")
    worksheet.write("F1", "点击次数")
    worksheet.write("G1", "直播间超过1分钟观看人次")
    worksheet.write("H1", "直播间查看购物车次数")
    worksheet.write("I1", "直播间商品点击次数")
    worksheet.write("J1", "直接成交金额(元)")
    worksheet.write("K1", "直接成交订单数")

    workbook.close()

    print(f"create {file} success :-)\n")
    logging.info(f"create {file} success!\n")
    return file

# 从配置文件中读取信息
def get_info_from_toml():
    info = toml.load("./config.toml")
    data_list = info.get('info')
    print(f"{data_list=}")
    return data_list

# 将数据写入文件中
def write_date_2_excel_file(file_name: str):
    year, month, day, hour = get_last_hour_time_info()

    # 初始化有关数据总和为0
    cost_sum = show_cnt_sum = click_cnt_sum = more_1_minute_cnt_sum = sidecar_click_cnt_sum = 0
    product_click_cnt_sum = direct_order_pay_gmv_sum = direct_order_pay_count_sum = 0

    today = get_today_str()

    user_info_list = get_info_from_toml()
    t_timestamp = int(get_last_hour_time_info())

    for user_info in user_info_list:
        name = user_info.get("name")
        vid = user_info.get("id")
        item = get_url_data(aavid=vid, today=today, user_info=user_info, time_info_timestamp=t_timestamp)

        if not item:
            print(f"{vid=} in the last hour does not exist data...")
            logging.info(f"{vid=} in the last hour does not exist data...")
            continue

        wb = load_workbook(filename=file_name)
        ws = wb.active
        current_row = ws.max_row + 1
        logging.info(f"{current_row=}\n")
        ws.cell(current_row, 1).value = f"{year}/{month}/{day}"
        ws.cell(current_row, 2).value = f"{hour}:00~{hour+1}:00"
        ws.cell(current_row, 3).value = name


        cost = item.get("cost")
        show_cnt = item.get("show_cnt")
        click_cnt = item.get("click_cnt")
        more_1_minute_cnt = item.get("more_1_minute_cnt")
        sidecar_click_cnt = item.get("sidecar_click_cnt")
        product_click_cnt = item.get("product_click_cnt")
        direct_order_pay_gmv = item.get("direct_order_pay_gmv")
        direct_order_pay_count = item.get("direct_order_pay_count")

        # need convert data
        cost = cost / 100000
        direct_order_pay_gmv = direct_order_pay_gmv / 100000


        ws.cell(current_row, 4).value = cost
        ws.cell(current_row, 5).value = show_cnt
        ws.cell(current_row, 6).value = click_cnt
        ws.cell(current_row, 7).value = more_1_minute_cnt
        ws.cell(current_row, 8).value = sidecar_click_cnt
        ws.cell(current_row, 9).value = product_click_cnt
        ws.cell(current_row, 10).value = direct_order_pay_gmv
        ws.cell(current_row, 11).value = direct_order_pay_count

        cost_sum += cost
        show_cnt_sum += show_cnt
        click_cnt_sum += click_cnt
        more_1_minute_cnt_sum += more_1_minute_cnt
        sidecar_click_cnt_sum += sidecar_click_cnt
        product_click_cnt_sum += product_click_cnt
        direct_order_pay_gmv_sum += direct_order_pay_gmv
        direct_order_pay_count_sum += direct_order_pay_count
        wb.save(file_name)
        logging.info("write single data to file\n")
        print("write single data to file\n")
        wb.close()

    wb = load_workbook(filename=file_name)
    ws = wb.active
    current_row = ws.max_row + 2

    ws.cell(current_row, 1).value = f"{year}/{month}/{day}"
    ws.cell(current_row, 2).value = f"{hour}:00~{hour+1}:00"
    ws.cell(current_row, 3).value = "总计"
    ws.cell(current_row, 4).value = cost_sum
    ws.cell(current_row, 5).value = show_cnt_sum
    ws.cell(current_row, 6).value = click_cnt_sum
    ws.cell(current_row, 7).value = more_1_minute_cnt_sum
    ws.cell(current_row, 8).value = sidecar_click_cnt_sum
    ws.cell(current_row, 9).value = product_click_cnt_sum
    ws.cell(current_row, 10).value = direct_order_pay_gmv_sum
    ws.cell(current_row, 11).value = direct_order_pay_count_sum
    wb.save(file_name)
    wb.close()
    logging.info("write sum data to file\n")
    print("write sum data to file\n")


# copy and rename file
def copy_and_rename(file: str):
    new_file = file.rstrip(".xlsx")+"-copy"+".xlsx"
    import shutil
    shutil.copy(file, new_file)
    logging.info(f"finish writing all data into file, please open {new_file} get the data you need, and wating 1 hour then get data :-) ")
    print(f"finish writing all data into file, please open {new_file} get the data you need, and wating 1 hour then get data :-) ")
    


"""
    for item in data_list:
        print(f"{current_row=}")
        ws.cell(current_row, 1).value = f"{year}/{month}/{day}"
        ws.cell(current_row, 2).value = f"{hour}:00~{hour+1}:00"
        ws.cell(current_row, 3).value = item.get('name')

        ws.cell(current_row, 4).value = item.get('cost')
        ws.cell(current_row, 5).value = item.get('show_cnt')
        ws.cell(current_row, 6).value = item.get('click_cnt')
        ws.cell(current_row, 7).value = item.get('more_1_minute_cnt')
        ws.cell(current_row, 8).value = item.get('sidecar_click_cnt')
        ws.cell(current_row, 9).value = item.get('product_click_cnt')
        ws.cell(current_row, 10).value = item.get('direct_order_pay_gmv')
        ws.cell(current_row, 11).value = item.get('direct_order_pay_count')

        cost_sum += item.get('cost')
        show_cnt_sum += item.get('show_cnt')
        click_cnt_sum += item.get('click_cnt')
        more_1_minute_cnt_sum += item.get('more_1_minute_cnt')
        sidecar_click_cnt_sum += item.get('sidecar_click_cnt')
        product_click_cnt_sum += item.get('product_click_cnt')
        direct_order_pay_gmv_sum += item.get('direct_order_pay_gmv')
        direct_order_pay_count_sum += item.get('direct_order_pay_count')

        current_row += 1
"""

# if __name__ == '__main__':
#     file_name = gen_xlsx()
#     write_date_2_excel_file(file_name)

# get_url_data(aavid=1729353507265543, today="2022-06-13")

"""
{
    "clickCnt": {
        "value": 91,
        "controlType": 1
        点击次数
        
    },
    "lubanLiveSlidecartClickCnt": {
        "value": 57,
        "controlType": 1
        直播间查看购物车次数
    },
    "lubanLivePayOrderGmv": {
        "value": 825000,
    },
    "lubanLivePayOrderCount": {
        "value": 1,
        直接成交订单数
    },
    "cpm": {
        "value": 3620336.087,
        "controlType": 1
        平均千次展现费用（元）36。20
    },
    "lubanLiveShareCnt": {
        "value": 1,
        直播间分享次数
    },
    "convertCost": {
        "value": 25422000,
        转换成本（元）：254。22
        
    },
    "videoOrderPayGmv": {
        "value": 0,
    },
    "directOrderPayGmv": {
        "value": 825000,
         直接成交金额（元）   8.25
    },
    "cost": {
        "value": 25422000,
        消耗（元）：  254。22元
    },
    "statCost": {
        "value": 25422000,
    },
    "showCnt": {
        "value": 7022,
        展示次数
    },
    "ctr": {
        "value": 1.296,
        点击率（%）      保留两位
    }
    "lubanLiveClickProductCnt": {
        "value": 50,
        直播间商店点击次数
    },
    "payOrderAmount": {
        "value": 825000,
    },
    "indirectOrderPayGmv7Days": {
        "value": 0,
        间接成交金额
    },
    "convertCnt": {
        "value": 1,
        转化数
    },
    "dyComment": {
        "value": 0,
        评论次数
    },
    "liveWatchOneMinuteCount": {
        "value": 12,
        直播间超过一分钟观看人数
    },
    "payOrderCount": {
        "value": 1,
    },
    "lubanLiveEnterCnt": {
        "value": 100,
        直播间观看人数1
    },
    "lubanLiveCommentCnt": {
        "value": 25,
        直播间评论次数
    },
    "dyFollow": {
        "value": 4,
        新增粉丝数
    },
    "videoOrderPayCount": {
        "value": 0,
    },
    "directOrderPayCount": {
        "value": 1,
        直接成交订单数
    },
    "convertRate": {
        "value": 1.099
        转换率 保留两位（四舍五入）1。10
    },
    "indirectOrderPayCount7Days": {
        "value": 0,
        间接订单成交数
    },
    "allOrderPayCount7Days": {
        "value": 1,
    },
    "allOrderPayGmv7Days": {
        "value": 825000,
    }
}
"""

"""
{
  "indirectOrderPayGmv7Days": {
    "value": 0,
    "controlType": 1
  },
  "showCnt": {
    "value": 840,
    "controlType": 1
  },
  "ctr": {
    "value": 0.8333333333333334,
    "controlType": 1
  },
  "lubanLivePayOrderCount": {
    "value": 0,
    "controlType": 1
  },
  "clickCnt": {
    "value": 7,
    "controlType": 1
  },
  "lubanLivePrepayOrderGmv": {
    "value": 0,
    "controlType": 1
  },
  "videoOrderPrepayDeductionGmv": {
    "value": 0,
    "controlType": 1
  },
  "videoOrderPrepayGmv": {
    "value": 0,
    "controlType": 1
  },
  "lubanLivePrepayOrderDeductionGmv": {
    "value": 0,
    "controlType": 1
  },
  "payOrderAmount": {
    "value": 0,
    "controlType": 1
  },
  "lubanLivePayOrderGmv": {
    "value": 0,
    "controlType": 1
  },
  "payOrderCount": {
    "value": 0,
    "controlType": 1
  },
  "videoOrderPayCount": {
    "value": 0,
    "controlType": 1
  },
  "cpm": {
    "value": 5845238.095238095,
    "controlType": 1
  },
  "allOrderPayRoi7Days": {
    "value": 0,
    "controlType": 1
  },
  "cost": {
    "value": 4910000,
    "controlType": 1
  },
  "videoOrderPayGmv": {
    "value": 0,
    "controlType": 1
  },
  "prepayAndPayOrderRoi": {
    "value": 0,
    "controlType": 1
  },
  "statCost": {
    "value": 4910000,
    "controlType": 1
  }
}


"""
